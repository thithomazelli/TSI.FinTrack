"""
Extrai dados dos xlsx anuais de orçamento e gera seed_data.json.

Uso:
  python3 scripts/parse-xlsx.py

Requer: pip install openpyxl msoffcrypto-tool
"""

import msoffcrypto, io, json, re, calendar, uuid
from pathlib import Path
import openpyxl
from datetime import date, datetime

PASSWORD    = "l#p01091910"
OWNER_ID    = "69f852bc-af5a-4f11-b293-37bf2f809018"
UPLOADS_DIR = Path(r"D:\Google Drive\Arquivos Thiago\Meus Gastos\Orçamentos encerrados")
OUT_FILE    = Path(__file__).parent.parent / "supabase" / "seed" / "transactions_data.json"

# ── Card metadata (name → due_day) ───────────────────────────────────────────
CARD_DUE = {
    "crédito nubank":            29,
    "crédito latam pass":        27,
    "crédito itaú multi pontos": 27,
    "crédito mastercard":        10,
    "crédito visa":              10,
}

# Normalise raw tipo string from spreadsheet → canonical card name
TIPO_MAP = {
    "débito":                          None,  # debit – no card
    "crédito mastercard":              "Crédito Mastercard",
    "cartão de crédito mastercard":    "Crédito Mastercard",
    "crédito visa":                    "Crédito Visa",
    "crédito nubank":                  "Crédito Nubank",
    "crédito latam pass":              "Crédito Latam Pass",
    "crédito itaú multi pontos":       "Crédito Itaú Multi Pontos",
}

# Rows to skip (header/summary labels in col B)
SKIP_B = {
    "resumo anual", "saldo atual", "renda mensal", "item", "total",
    "despesas mensais", "tipo", "porcentagem da renda gasta",
    "rótulos de linha", "resumo", "grand total",
}

MONTH_RE = re.compile(r"^(\d{2})\.")  # "01. Janeiro" → group 1 = "01"

# "Descrição 3/12" → (base, installment_num, total)
INSTALLMENT_RE = re.compile(r"^(.*?)\s+(\d{1,2})/(\d{1,2})$")

def parse_installment(description: str):
    """Return (base, num, total) if description matches 'Base X/Y', else None."""
    m = INSTALLMENT_RE.match(description.strip())
    if not m:
        return None
    num, total = int(m.group(2)), int(m.group(3))
    if num < 1 or total < 2 or num > total:
        return None
    if total > 120:  # >120 months is not a real installment (likely MM/YYYY pattern)
        return None
    return m.group(1).strip(), num, total

# Normalise category names that changed over the years
CAT_ALIAS = {
    "jogos":                  "Games",
    "uber / 99":              "Uber/99",
    "cuidados pessoais":      "Cuidados Pessoais",
    "despesas carro":         "Despesas carro",
    "despesas casa":          "Despesas casa",
    "despesas empresa":       "Despesas Empresa",
    "despesas terreno":       "Despesas terreno",
    "saude":                  "Saúde",
    "saúde":                  "Saúde",
    "vestuario":              "Vestuário",
    "vestuário":              "Vestuário",
    "eletronicos":            "Eletrônicos",
    "eletrônicos":            "Eletrônicos",
    "poupanca":               "Poupança",
    "poupança":               "Poupança",
    # case variants
    "alimentação/mercado":    "Alimentação/Mercado",
    "empréstimo pessoal":     "Empréstimo Pessoal",
    "emprestimo pessoal":     "Empréstimo Pessoal",
    "empréstimo bancário":    "Empréstimo bancário",
    "emprestimo bancario":    "Empréstimo bancário",
    "telefone celular":       "Celular",
    "ti":                     "TI",
    "transporte público":     "Transporte Público",
    "transporte publico":     "Transporte Público",
    "farmacia":               "Farmácia",
    "farmácia":               "Farmácia",
    "medico":                 "Médico",
    "médico":                 "Médico",
    "tenis":                  "Tênis",
    "tênis":                  "Tênis",
    "viagem":                 "Viagem",
    "roupas":                 "Roupas",
    "parcela carro":          "Parcela carro",
}

# ── helpers ──────────────────────────────────────────────────────────────────

def open_xlsx(path: Path):
    with open(path, "rb") as f:
        try:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=PASSWORD)
            buf = io.BytesIO()
            office.decrypt(buf)
        except Exception:
            f.seek(0)
            buf = io.BytesIO(f.read())
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


def to_date_str(val, fallback_year: int, fallback_month: int, enforce_month: bool = False) -> str | None:
    """Convert a cell value to an ISO date string (YYYY-MM-DD).

    Handles: datetime/date objects from openpyxl, text strings (M/D/YYYY or
    YYYY-MM-DD), and Excel float serial numbers.
    enforce_month=True: force the result into fallback_year/fallback_month.
    """
    d = None

    if isinstance(val, datetime):
        d = val.date()
    elif isinstance(val, date):
        d = val
    elif isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ('%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%y'):
            try:
                d = datetime.strptime(val, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            return None
    elif isinstance(val, (float, int)):
        # Excel date serial number (openpyxl may return int for whole-number dates)
        try:
            epoch = date(1899, 12, 30)
            from datetime import timedelta
            d = epoch + timedelta(days=int(val))
        except Exception:
            return None
    else:
        return None

    # Dates before 2000 almost certainly have the wrong year (cell stores only day).
    if d.year < 2000:
        last = calendar.monthrange(fallback_year, fallback_month)[1]
        return date(fallback_year, fallback_month, min(d.day, last)).isoformat()
    if enforce_month and (d.year != fallback_year or d.month != fallback_month):
        last = calendar.monthrange(fallback_year, fallback_month)[1]
        return date(fallback_year, fallback_month, min(d.day, last)).isoformat()
    return d.isoformat()


def coerce_amount(val):
    """Return a float amount from val, handling cells mis-typed as datetime by Excel."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, (datetime, date)):
        # The cell has a numeric value stored with a date format; convert back to serial.
        d = val.date() if isinstance(val, datetime) else val
        t = (val.hour*3600 + val.minute*60 + val.second + val.microsecond/1e6
             if isinstance(val, datetime) else 0)
        days = (d - date(1900, 1, 1)).days + 1
        if d >= date(1900, 3, 1):  # Excel's phantom 1900-02-29 (serial 60)
            days += 1
        return days + t / 86400
    return None


def payment_date(year: int, month: int, due_day: int) -> str:
    """Return YYYY-MM-DD for due_day clamped to last day of month."""
    last = calendar.monthrange(year, month)[1]
    day = min(due_day, last)
    return date(year, month, day).isoformat()


# ── parser ───────────────────────────────────────────────────────────────────

# Months after which 2026 entries/transactions become PROJECTED
PROJECTED_FROM = (2026, 7)  # July 2026 onwards

# Savings movements before this year are ignored (data histórica incompleta)
SAVINGS_MIN_YEAR = 2015

def row_status(year: int, month: int) -> str:
    if (year, month) >= PROJECTED_FROM:
        return "PROJECTED"
    return "REALIZED"


def parse_sheet(ws, year: int, month: int):
    entries      = []
    transactions = []
    savings      = []
    status       = row_status(year, month)

    in_income   = False
    in_expense  = False

    for row_cells in ws.iter_rows():
        row_num = row_cells[0].row
        raw_row = tuple(c.value for c in row_cells)
        b = raw_row[1]  # col B

        # Section headers
        if isinstance(b, str):
            bl = b.strip().lower()
            if "renda mensal" in bl:
                in_income  = True
                in_expense = False
                continue
            if "despesas mensais" in bl:
                in_income  = False
                in_expense = True
                continue
            # Fallback: detect expense section from its column header row
            # (some months are missing the "DESPESAS MENSAIS" label)
            if bl == "tipo" and isinstance(raw_row[2], str) and raw_row[2].strip().lower() == "item":
                in_income  = False
                in_expense = True
                continue
            if bl == "total" and in_income:
                in_income = False
                continue
            if bl == "total" and in_expense:
                in_expense = False
                continue

        # ── Income row ────────────────────────────────────────────────────
        if in_income:
            item    = b                 # col B = item/description
            dt_val  = raw_row[2]        # col C = date
            amount  = raw_row[3]        # col D = amount

            if not item or not isinstance(item, str):
                continue
            if item.strip().lower() in SKIP_B:
                continue
            amount = coerce_amount(amount)
            if not amount:
                continue

            dt_str = to_date_str(dt_val, year, month, enforce_month=True) or date(year, month, calendar.monthrange(year, month)[1]).isoformat()

            entries.append({
                "owner_id":    OWNER_ID,
                "description": item.strip(),
                "amount":      round(float(amount), 2),
                "date":        dt_str,
                "status":      status,
                "labels":      [],
                "position":    row_num,
            })

            # Renda com item == "Poupança" → renomeia descrição e registra retirada
            if item.strip().lower() == "poupança":
                entries[-1]["description"] = "Resgate Poupança"
                if year >= SAVINGS_MIN_YEAR:
                    savings.append({
                        "owner_id":    OWNER_ID,
                        "description": "Resgate Poupança",
                        "amount":      round(abs(float(amount)), 2),
                        "date":        dt_str,
                        "type":        "WITHDRAWAL",
                    })

        # ── Expense row ───────────────────────────────────────────────────
        if in_expense:
            tipo    = b                 # col B = tipo (card / Débito)
            cat     = raw_row[2]        # col C = category
            dt_val  = raw_row[3]        # col D = date (purchase date)
            desc    = raw_row[4]        # col E = description
            amount  = raw_row[5]        # col F = amount

            if not tipo or not isinstance(tipo, str):
                continue
            tipo_l = tipo.strip().lower()
            if tipo_l in SKIP_B:
                continue
            if tipo_l not in TIPO_MAP:
                continue
            amount = coerce_amount(amount)
            if not amount:
                continue

            card_name = TIPO_MAP[tipo_l]   # None = debit

            if card_name:
                # Credit: purchase date from col D (same as debit uses col D for date)
                purchase_dt = to_date_str(dt_val, year, month)
                due_day = CARD_DUE.get(card_name.lower(), 10)
                bill_dt  = payment_date(year, month, due_day)
                # Fallback: if col D is empty/unparseable, use last day of billing month
                if not purchase_dt:
                    last_day = calendar.monthrange(year, month)[1]
                    purchase_dt = date(year, month, last_day).isoformat()
                purch = purchase_dt
            else:
                # Debit: date must belong to the sheet's month — enforce it
                purchase_dt = to_date_str(dt_val, year, month, enforce_month=True)
                bill_dt  = purchase_dt or date(year, month, calendar.monthrange(year, month)[1]).isoformat()
                purch    = None

            cat_str = (cat or "").strip() if isinstance(cat, str) else ""
            cat_str = CAT_ALIAS.get(cat_str.lower(), cat_str)

            transactions.append({
                "owner_id":         OWNER_ID,
                "description":      (desc or cat or tipo).strip(),
                "amount":           round(float(amount), 2),
                "date":             bill_dt,
                "purchase_date":    purch,
                "category_name":    cat_str,
                "credit_card_name": card_name,
                "status":           status,
                "labels":           [],
                "position":         row_num,
            })

    # ── Savings: despesas categoria "Poupança" → DEPOSIT (apenas >= 2015) ───
    if year >= SAVINGS_MIN_YEAR:
        for t in transactions:
            if (t.get("category_name") or "").lower() != "poupança":
                continue
            savings.append({
                "owner_id":    OWNER_ID,
                "description": t["description"],
                "amount":      round(abs(t["amount"]), 2),
                "date":        t.get("purchase_date") or t["date"],
                "type":        "DEPOSIT",
            })

    return entries, transactions, savings


def parse_workbook(path: Path, year: int):
    wb = open_xlsx(path)
    all_entries      = []
    all_transactions = []
    all_savings      = []

    for sheet_name in wb.sheetnames:
        m = MONTH_RE.match(sheet_name)
        if not m:
            continue
        month = int(m.group(1))
        ws = wb[sheet_name]
        e, t, s = parse_sheet(ws, year, month)
        yr_str = str(year)
        # Only keep records whose date belongs to this workbook's year to avoid
        # duplicates when a previous year's file contains projected future months.
        e = [x for x in e if x["date"].startswith(yr_str)]
        t = [x for x in t if x["date"].startswith(yr_str)]
        all_entries.extend(e)
        all_transactions.extend(t)
        all_savings.extend(s)
        dep  = sum(1 for x in s if x["type"] == "DEPOSIT")
        wdw  = sum(1 for x in s if x["type"] == "WITHDRAWAL")
        print(f"  {year}/{month:02d} {sheet_name}: {len(e)} entries, {len(t)} transactions, {len(s)} savings ({dep}D/{wdw}W)")

    return all_entries, all_transactions, all_savings


# ── Global installment grouping ───────────────────────────────────────────────

def assign_installment_groups(all_transactions: list) -> int:
    """
    Scan all transactions globally (across all months/years), group those whose
    description matches 'Base X/Y', and assign a shared installment_group_id.
    Must run before generate_pending_installments.
    """
    buckets: dict[tuple, list] = {}
    for t in all_transactions:
        parsed = parse_installment(t["description"])
        if not parsed:
            continue
        base, num, total = parsed
        t["installment_number"] = num
        t["total_installments"] = total
        key = (base.lower(), t.get("credit_card_name") or "", total)
        buckets.setdefault(key, []).append((t, num))

    groups_created = 0
    for (base_key, card_key, total), items in buckets.items():
        # Sort by date then installment number
        items.sort(key=lambda x: (x[0]["date"], x[1]))
        for chunk_start in range(0, len(items), total):
            chunk = items[chunk_start: chunk_start + total]
            if len(chunk) < 2:
                continue
            group_id = str(uuid.uuid4())
            for t, num in chunk:
                t["installment_group_id"] = group_id
            groups_created += 1
    return groups_created


# ── Installment continuation ──────────────────────────────────────────────────

def generate_pending_installments(all_transactions: list) -> list:
    """
    For each installment group where the highest seen installment < total,
    generate the missing future installments as PROJECTED transactions.

    The billing date for each missing installment is computed by advancing
    the last seen billing date by 1 month per installment.
    """
    # Group by installment_group_id
    groups: dict[str, list] = {}
    for t in all_transactions:
        gid = t.get("installment_group_id")
        if not gid:
            continue
        groups.setdefault(gid, []).append(t)

    generated = []
    for gid, items in groups.items():
        items_sorted = sorted(items, key=lambda x: x.get("installment_number", 0))
        last = items_sorted[-1]
        total = last.get("total_installments", 0)
        last_num = last.get("installment_number", 0)

        if last_num >= total:
            continue  # all installments already present

        # Parse base description from last item
        parsed = parse_installment(last["description"])
        if not parsed:
            continue
        base_desc, _, _ = parsed

        # Compute the month after the last billing date
        last_date = date.fromisoformat(last["date"][:10])
        # Use purchase_date of last known installment as base for projecting purchase dates
        last_purch_raw = last.get("purchase_date")
        last_purch = date.fromisoformat(last_purch_raw[:10]) if last_purch_raw else last_date
        card_name = last.get("credit_card_name")
        due_day   = CARD_DUE.get((card_name or "").lower(), 10) if card_name else None

        for n in range(last_num + 1, total + 1):
            # Advance month
            delta = n - last_num
            m = last_date.month + delta
            y = last_date.year + (m - 1) // 12
            m = ((m - 1) % 12) + 1
            last_day = calendar.monthrange(y, m)[1]

            if card_name and due_day:
                bill_dt = date(y, m, min(due_day, last_day)).isoformat()
            else:
                bill_dt = date(y, m, last_day).isoformat()

            # Advance purchase date by same number of months
            pm = last_purch.month + delta
            py = last_purch.year + (pm - 1) // 12
            pm = ((pm - 1) % 12) + 1
            purch_dt = date(py, pm, min(last_purch.day, calendar.monthrange(py, pm)[1])).isoformat()

            new_t = {**last}
            new_t["description"]         = f"{base_desc} {n}/{total}"
            new_t["date"]                = bill_dt
            new_t["purchase_date"]       = purch_dt
            new_t["installment_number"]  = n
            new_t["status"]              = "PROJECTED"
            new_t["position"]            = None
            generated.append(new_t)

    return generated


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    # Find all xlsx files with a 4-digit year anywhere in the name
    year_files: dict[int, Path] = {}
    for f in sorted(UPLOADS_DIR.glob("*.xlsx")):
        if f.name.startswith("~$"): continue  # skip Excel temp/lock files
        m = re.search(r"(\d{4})", f.name)
        if not m: continue
        year = int(m.group(1))
        if year < 2000 or year > 2100: continue
        if year not in year_files or f.stat().st_mtime > year_files[year].stat().st_mtime:
            year_files[year] = f

    print(f"Encontrados {len(year_files)} arquivos:")
    for y, p in sorted(year_files.items()):
        print(f"  {y} -> {p.name}")
    print()

    all_entries      = []
    all_transactions = []
    all_savings      = []

    for year in sorted(year_files):
        path = year_files[year]
        print(f"Processando {year} ({path.name})…")
        e, t, s = parse_workbook(path, year)
        all_entries.extend(e)
        all_transactions.extend(t)
        all_savings.extend(s)
        print(f"  -> subtotal: {len(e)} entries, {len(t)} transactions, {len(s)} savings\n")

    # ── Global installment grouping (across all months/years) ────────────────
    groups = assign_installment_groups(all_transactions)
    print(f"Grupos de parcelas detectados: {groups}")

    # ── Installment continuation: project remaining parcelas beyond last year ──
    projected = generate_pending_installments(all_transactions)
    if projected:
        print(f"Parcelas projetadas geradas: {len(projected)}")
        for t in projected:
            parsed = parse_installment(t["description"])
            print(f"  {t['date']}  {t['description']}  R$ {t['amount']:.2f}  [{t.get('credit_card_name') or 'debito'}]")
        all_transactions.extend(projected)

    result = {
        "meta": {"opening_balance": -305, "opened_at": "2009-05-01"},
        "entries":      all_entries,
        "transactions": all_transactions,
        "savings":      all_savings,
    }

    total_dep_n   = sum(1 for s in all_savings if s["type"] == "DEPOSIT")
    total_wdw_n   = sum(1 for s in all_savings if s["type"] == "WITHDRAWAL")
    total_dep_amt = sum(s["amount"] for s in all_savings if s["type"] == "DEPOSIT")
    total_wdw_amt = sum(s["amount"] for s in all_savings if s["type"] == "WITHDRAWAL")
    OUT_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\nOK: {len(all_entries)} entries + {len(all_transactions)} transactions")
    print(f"   savings: {len(all_savings)} total")
    print(f"   DEPOSIT  : {total_dep_n:4d} movimentos  R$ {total_dep_amt:,.2f}")
    print(f"   WITHDRAWAL: {total_wdw_n:4d} movimentos  R$ {total_wdw_amt:,.2f}")
    print(f"   SALDO     :                  R$ {total_dep_amt - total_wdw_amt:,.2f}")
    print(f"   -> {OUT_FILE}")


if __name__ == "__main__":
    main()
